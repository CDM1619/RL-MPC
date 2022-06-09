# -*- coding: utf-8 -*-
"""
Created on Sun Jan 16 13:33:03 2022

@author: 86153
"""

import pprint
from cvxopt import matrix, solvers
# import sys 
# sys.path.append('C:\\Users\\86153\\Desktop\\CarCombat_test_2\\CarCombat_fight_test_2')
from openpyxl import Workbook
from math import sqrt, pow
import torch
import os
import numpy as np
import random
import gym
import math
import matplotlib.pyplot as plt
from matplotlib.pyplot import MultipleLocator
import warnings
warnings.filterwarnings("ignore")
from matplotlib.patches import Circle
import copy
import imageio,os
from matplotlib.animation import FuncAnimation

from SAC import config,SAC
from car_mpc import CarCombatEnv
# from CarCombat_dyna_avoid_test import CarCombatEnv


class MPC_QP_solver():
    def __init__(self):
        return
        
        

    def sol(self,x,y,pusin,v,x_r,y_r,v_r,pusin_r,l_r,l_f,T,a_max,delta_max,a_min,delta_min,N_p):  #T是采样周期
 
        # print([x,y,pusin,v,x_r,y_r,pusin_r])
        # delta_r = np.pi/120
        # delta = (action * action_mean + action_bias)[1]
        # delta = action_MPC[1]
        
        # belta = math.atan((l_r / (l_r + l_f)) * math.tan(delta))
        # belta_r = math.atan((l_r / (l_r + l_f)) * math.tan(delta_r))
        delta_r = 0
       #计算A矩阵  
        A = 0
        A = copy.deepcopy(np.array([[1, 0, -v_r * T * math.sin(pusin_r), T * math.cos(pusin_r)],
                      [0, 1, v_r * T * math.cos(pusin_r), T * math.sin(pusin_r)],
                      [0, 0, 1, T * math.tan(delta_r) / (l_r + l_f)],
                      [0, 0, 0, 1]]))
        
        
        #计算B矩阵
        B = 0
        B = copy.deepcopy(np.array([[0, 0],
                      [0, 0],
                      [0, (v_r *T) / ((l_r + l_f)*math.cos(delta_r)**2) ],
                      [T, 0]]))
        
        #计算E矩阵
        E = 0
        E = copy.deepcopy(np.array([[x - x_r],
                      [y - y_r],
                      [pusin - pusin_r],
                      [v - v_r]]))
        
        #计算系数矩阵Q
        Q = 0
        q=copy.deepcopy([5, 5, 0, 5]) * N_p   #从左至右分别表示每一项状态量误差优化的权重（x轴距离偏差，y轴距离偏差，偏航角pusin偏差，速度v偏差）
        Q = np.diag(copy.deepcopy(q))
        
        
        #计算系数矩阵R
        R = 0
        r = copy.deepcopy([0, 0]) * N_p     #从左至右分别表示每一项动作量优化的权重（加速度a，转角delta）
        R = np.diag(copy.deepcopy(r))
        
        #计算G矩阵
        G = 0
        G=copy.deepcopy(A)
        a=copy.deepcopy(A)
        for i_1 in range(N_p-1):
            a=np.dot(copy.deepcopy(A),copy.deepcopy(a))
            G = np.vstack((copy.deepcopy(G), copy.deepcopy(a)))
        # print(G)
        
        
        #构造函数，计算矩阵M的n次方
        def matrix_n_power(M,n):
            if n == 0:
                return np.eye(np.shape(M)[0])
            elif n == 1:
                return copy.deepcopy(M)
            a=copy.deepcopy(M)
            for i_2 in range(n-1):
                a=np.dot(copy.deepcopy(M),copy.deepcopy(a))
            return a
        
        
        #构造函数，计算H矩阵的每一列构成,n表示第几列
        def H_column(A,B,n,N_p):   
            b=copy.deepcopy(B)
            z=np.zeros((np.shape(B)[0], np.shape(B)[1])) 
            if n == N_p-1:
                b = np.vstack((np.zeros((n*np.shape(B)[0], np.shape(B)[1])) ,copy.deepcopy(b)))
                
            elif n == 0:
                for i_3 in range(N_p-1):
                    b = np.vstack((copy.deepcopy(b), np.dot(matrix_n_power(copy.deepcopy(A),i_3+1),copy.deepcopy(B))))
        
            else:
                for i_4 in range(N_p-(n+1)):
                    b = np.vstack((copy.deepcopy(b), np.dot(matrix_n_power(copy.deepcopy(A),i_4+1),copy.deepcopy(B))))
                z = np.zeros((n*np.shape(B)[0], np.shape(B)[1])) 
                b = np.vstack((copy.deepcopy(z),copy.deepcopy(b)))
            return b
        
        # print(N_p)        
                    
        
        #计算H矩阵
        H = 0
        H = copy.deepcopy(H_column(copy.deepcopy(A),copy.deepcopy(B),0,N_p))
        for i_5 in range(N_p-1):
            H = np.hstack((copy.deepcopy(H), copy.deepcopy(H_column(copy.deepcopy(A),copy.deepcopy(B),i_5+1,N_p))))
            
        
        #计算cvxopt求解器中的矩阵：result = solvers.qp(P,q,G,h)
            
        #计算P矩阵
        P_cvx = 0
        q_cvx = 0
        G_cvx = 0
        h_cvx = 0        
        
        P_cvx = 2 * ((np.dot(np.dot((copy.deepcopy(H)).transpose(),copy.deepcopy(Q)),copy.deepcopy(H))) + copy.deepcopy(R))
        q_cvx = 2 * np.dot(np.dot(np.dot((copy.deepcopy(H)).transpose(), copy.deepcopy(Q).transpose()),copy.deepcopy(G)),copy.deepcopy(E))
        G_cvx = np.vstack((np.eye(2*N_p), -np.eye(2*N_p)))
        
        # print(G_cvx)        

        
        action_max = np.array([[a_max],[delta_max]])
        action_min = np.array([[a_min],[delta_min]])
        # print(action_min)
        h_1 = copy.deepcopy(action_max) 
        for i_6 in range(N_p-1):
            h_1 = np.vstack((copy.deepcopy(h_1), copy.deepcopy(action_max)))
        h_2 = copy.deepcopy(-action_min) 
        for i_7 in range(N_p-1):
            h_2 = np.vstack((copy.deepcopy(h_2), copy.deepcopy(-action_min)))
        
        # print(h_2)
            
        h_cvx = np.vstack((copy.deepcopy(h_1), copy.deepcopy(h_2)))
        P_cvx_ = matrix(copy.deepcopy(P_cvx))
        q_cvx_ = matrix(copy.deepcopy(q_cvx))
        G_cvx_ = matrix(copy.deepcopy(G_cvx))
        h_cvx_ = matrix(copy.deepcopy(h_cvx))
        

        
        result = 0
        result = copy.deepcopy(solvers.qp(copy.deepcopy(P_cvx_),copy.deepcopy(q_cvx_),copy.deepcopy(G_cvx_),copy.deepcopy(h_cvx_)))
        
        
        return result['x']
                    
        # return result




def run(env, agent):
    E =[]
    R = []
    R_b = []
    
    R__ = []
    R_b__ = []  
    R_error=[]
    R_0=[]
    
    
    #用来计算N回合中平均胜率
    Win_Total_A=[]    
    Win_Total_B=[]   
    Episode_Adv_A=[]
    Episode_Adv_B=[]  
    
    # X = []
    # for i in range(1000):
    #     X.append(i)

    # Y = [0]*1000
    for episode in range(Episode):
        E.append(episode)
        x_a = []
        y_a = []
        x_b = []
        y_b = []
        
        V_a = []
        V_b = []
        
        D = []
        Angle_a = []
        Angle_b = []
        Step = []

        step_r_a=[]
        step_r_b=[]
        step_r_error=[]
        step_r_0=[]
        
        Line_d = []

        Line_zero = []
        Line_a = []
        score = 0
        score_b = 0
        
        Episode_advantage_a = 0
        Episode_advantage_b = 0
        
        
        

        p1_a=0
        p2_a=0
        h_a=0
        
        

        p1_b=0
        p2_b=0 
        h_b=0
        
        
        adv_a=0
        adv_b=0
        adv_balance=0

        error_1=0
        error_2=0        
        error_3=0


        
        u=0
        q=0
        
        state = env.reset()

        action_mean = (env.action_space.high - env.action_space.low) / 2
        action_bias = (env.action_space.high + env.action_space.low) / 2
        state_mean = (env.observation_space.high - env.observation_space.low)/2
        state_bias = (env.observation_space.high + env.observation_space.low)/2    
        
        # B车采取的策略：用A车之前训练好的策略网络      
        sac_b = SAC(env)
        sac_b.load_nets('./',19999)
        
        for i in range(step):
            state_norm = (state - state_bias) / state_mean         
            action_sac = agent.get_action_test(state_norm)   

            
            # MPC控制——————————————————————————————————————————————————————————————
            x = state[0]
            y = state[1]
            pusin = state[2]
            v = state[3]
            
            # x_r = 100
            # y_r = 100
            # pusin_r = 0
            # v_r = 20

            # x_r = state[4]
            # y_r = state[5]  
            x_r = state[4] - 12.5*math.cos(state[6])
            y_r = state[5] - 12.5*math.sin(state[6])

            pusin_r = state[6]
            v_r = state[7]
            
            
            l_r = l_f = 2.5
            T = 0.1
            a_max = 10
            delta_max = np.pi/4
            a_min = -10
            delta_min = -np.pi/4 
            N_p=10
            
            
            action_MPC = 0
            s_mpc = MPC_QP_solver()
            action_MPC = s_mpc.sol(x,y,pusin,v,x_r,y_r,v_r,pusin_r,l_r,l_f,T,a_max,delta_max,a_min,delta_min,N_p)
            
            # print(action_MPC)
            action = [copy.deepcopy(action_MPC[0]),copy.deepcopy(action_MPC[1])]
            # print(action)
            # print(i)
            action_mpc = (action - action_bias)/action_mean
            
            # if i>=300:
            #     action = [action_MPC[0],0]
            # ————————-—————————————————————————————————————————————————————————————— 
            
            
            alpha_1 = 1
            alpha_2 = 0
            action = (alpha_1 * action_sac + alpha_2 * action_mpc)
            # print(action)
            







            
            
            
            
            
            
            
            
            
            


            # #追踪和避障用两个网络  
            
            # min_m = env.get_min_m(state[0],state[1],state[2],state[3])
            # if min_m <= env.get_r()+10:
            #     agent = SAC(env)
            #     agent.load_nets('./',89999)
            #     action = agent.get_action_test(state)
            # else:
            #     agent = SAC(env)
            #     agent.load_nets('./',29999)
            #     action = agent.get_action_test(state)
            
            #追踪和避障用一个网络              
            # action = agent.get_action_test(state)
            
               
            # #让敌方动作选择按照已经训练好的网络进行
            # s_1 = state[4:8]
            # s_2 = state[0:4]
            # s_3 = state[8:9]
            
            # a=np.hstack((s_1,s_2))            
            # b=np.hstack((a,s_3))            
            # s_4 = [env.get_min_m(state[4],state[5],state[6],state[7])]
            # state_2=np.hstack((b,s_4))  
            # action_2 = sac_b.get_action_test(state_2)      
            action_2=[0,0]
            next_state, reward, done, _, o_x, o_y, reward__ = env.step(action * action_mean + action_bias, action_2 * action_mean + action_bias,episode,i,)   
            reward_b__ = env.get_b_step_reward(state[0], state[1], state[2], state[3], state[4], state[5], state[6], state[7], o_x, o_y)             
            
            # if i ==step-1:
            #     print(o_x)          
            
            
            # #让敌方动作选择按照随机给出
            # next_state, reward, done, _, o_x, o_y = env.step(action * action_mean + action_bias, 0,episode)
                  
            if done:
                break            
            d = env.get_distance()
            angle_1, angle_2 = env.get_angle()
            v_a = env.get_v_a()
            v_b = env.get_v_b()
            # o_x, o_y = env.get_obstacle() 
            
            q+=1


            

            state = next_state
            
            
            score = score + reward
            
            
            Episode_advantage_a += reward__
            Episode_advantage_b += reward_b__            
            #A车和B车每回合中每一步的优势数值比较：
            if reward__>reward_b__:        
                adv_a+=1
            elif reward__<reward_b__:        
                adv_b+=1
            elif reward__==reward_b__:        
                adv_balance+=1  
            if adv_a + adv_b!=0:                
                episode_adv_a = adv_a/(adv_a + adv_b)
                episode_adv_b = adv_b/(adv_a + adv_b)                
                episode_adv_balance = 1 - (adv_a + adv_b) 
            else:
                episode_adv_a = 0
                episode_adv_b = 0               
                episode_adv_balance = 1               
                 
                
            
           
            # A车优势比率计算
            if d<=20:         
                
                # 综合优势
                if d<=20 and angle_1<=50:
                    p1_a += 1
                else :
                    p2_a += 1

                    
            
                win_total_a = (p1_a)/(p1_a + p2_a)
                          #总体均势                            
                
            elif d>20:
                h_a += 1
                if h_a == q:                   
                    win_total_a = 0
                

            # B车优势比率计算
            if d<=20:
          
                # 综合优势
                if d<=20 and angle_2<=50 :
                    p1_b += 1
                else :
                    p2_b += 1

      
                win_total_b = (p1_b)/(p1_b + p2_b)
                              
                
            elif d>20:
                h_b += 1
                if h_b == q:
                    win_total_b = 0   
                    
                    
            if (reward__-reward_b__)>0:
                error_1+=1
            elif (reward__-reward_b__)<0:
                error_2+=1          
            elif (reward__-reward_b__)==0:
                error_3+=1
            win_error_1 = error_1/(error_1+error_2+error_3)
            win_error_2 = error_2/(error_1+error_2+error_3)
            win_error_3 = error_3/(error_1+error_2+error_3)            
            
            
            x_a.append(state[0])
            y_a.append(state[1])
            x_b.append(state[4])
            y_b.append(state[5])
            V_a.append(v_a)
            V_b.append(v_b)            
            D.append(d)
            Angle_a.append(angle_1)
            Angle_b.append(angle_2)
            Step.append(i)   
            Line_d.append(20)

            Line_zero.append(0)
            Line_a.append(50)
            step_r_a.append(reward__)
            step_r_b.append(reward_b__)  
            step_r_error.append((reward__-reward_b__)/10**3)
            step_r_0.append(0)


            if d <= 30:
                u+=1



        Episode_Adv_A.append(episode_adv_a)
        Episode_Adv_B.append(episode_adv_b)
            
        #用来计算N回合中A车和B车的平均优势比率 

        if (win_total_b + win_total_a)!=0:
            
            Win_Total_A.append((win_total_a/(win_total_b + win_total_a)*100))                         
            Win_Total_B.append((win_total_b/(win_total_b + win_total_a)*100))
        N=[]
        Line_c=[]
        Line_100=[]
        for i in range(1,len(Win_Total_A)+1):
            N.append(i)
            Line_c.append(50)
            Line_100.append(100)

 

        
        
          #绘制静态轨迹  (有障碍物)           
        if episode+1 == Episode:
            
            #新建一个文件夹用来存放数据          
            path = "./graph/paper/Good Result/"
            # 定义文件夹名称
            name = "test_result_"         
            i_=0 
            
            while True:
                
                isExists = os.path.exists(path+name+str(i_))
                isExists_next = os.path.exists(path+name+str(i_+1))
                
                if not isExists:
                    os.makedirs(path+name+str(i_)) 
                    j_=i_                  
                    break
                elif isExists and not isExists_next:
                    os.makedirs(path+name+str(i_+1))
                    j_=i_+1
                    break
                elif isExists and isExists:
                    i_+=1
                    continue
                                    
                 
            # plt.figure(1)
            # o_x, o_y = env.get_obstacle() 
            # plt.plot(o_x,o_y,'o')
            
            #plt.scatter(o_x, o_y, s=585, c='g')
            
            
            fig,ax=plt.subplots() 
            for j in range(0,len(o_x)):
                
                  x_r, y_r = (o_x[j],o_y[j])
                
                  cir1 = Circle(xy = (x_r, y_r), radius = env.get_r(), facecolor= 'black', zorder=j+20)
                  ax.add_patch(cir1)

                  plt.axis('scaled')
                  ax.axis('equal') 
            



            # #画三角形用来表示小车方向
            # kk = plt.Polygon(xy=[[100, 100], [100, 300], [300, 300]], color='red',zorder=14)
            # ax.add_patch(kk)
            
            
          
            
          
            
           #画矩形用来画小车           
            def angle_of_vector(v1, v2):
                pi = np.pi
                vector_prod = v1[0] * v2[0] + v1[1] * v2[1]
                length_prod = sqrt(pow(v1[0], 2) + pow(v1[1], 2)) * sqrt(pow(v2[0], 2) + pow(v2[1], 2))
                cos = vector_prod * 1.0 / (length_prod * 1.0 + 1e-6)
                return (math.acos(cos) / pi) * 180            


          #   # 方案1：车身长度与小车前后的步数长度相关  
          #   for i in range(1,4):
          #       l_x=len(x_a)
          #       x1_a = x_a[l_x//4*i]
          #       y1_a = y_a[l_x//4*i]
                
          #       x2_a = x_a[l_x//4*i+10]
          #       y2_a = y_a[l_x//4*i+10]
                           
                
          #       ang_a = angle_of_vector([x2_a-x1_a,y2_a-y1_a],[1,0])
    
                
          #       if x2_a-x1_a<0 and y2_a-y1_a<=0:
          #           ang_aa = -ang_a
          #           ang_a = 180-ang_a 
                    
          #       elif x2_a-x1_a<=0 and y2_a-y1_a>0:
          #           ang_aa = ang_a
          #           ang_a = ang_a-180 
    
                    
          #       elif x2_a-x1_a>0 and y2_a-y1_a>=0:
          #           ang_aa = ang_a
          #           ang_a = ang_a-180 
    
                    
          #       elif x2_a-x1_a>=0 and y2_a-y1_a<0:
          #           ang_aa = -ang_a
          #           ang_a = 180-ang_a    
                
                          
    
          #       # #车身(长度用前后步数差的距离衡量)                          
          #       # square1 = plt.Rectangle(xy=(x2_a, y2_a), width = math.sqrt((x1_a-x2_a)**2+(y1_a-y2_a)**2), height = 10, angle=ang_a, color='royalblue',zorder=8)                  
               
          #       #车身 （长度固定）
          #       square1 = plt.Rectangle(xy=(x2_a, y2_a), width = math.sqrt((x1_a-x2_a)**2+(y1_a-y2_a)**2), height = 10, angle=ang_a, color='royalblue',zorder=8)                  
                              
          #       #车身
          #       square2 = plt.Rectangle(xy=(x2_a, y2_a), width = math.sqrt((x1_a-x2_a)**2+(y1_a-y2_a)**2), height = -10, angle=ang_a, color='royalblue',zorder=8)                  
    
          #       #前车轮                          
          #       square3 = plt.Rectangle(xy=(x2_a, y2_a), width = math.sqrt((x1_a-x2_a)**2+(y1_a-y2_a)**2)/5, height = 15, angle=ang_a, color='black',zorder=7)                  
               
          #       #前车轮
          #       square4 = plt.Rectangle(xy=(x2_a, y2_a), width = math.sqrt((x1_a-x2_a)**2+(y1_a-y2_a)**2)/5, height = -15, angle=ang_a, color='black',zorder=7)                  
     
          #       #后车轮                          
          #       square5 = plt.Rectangle(xy=(x1_a, y1_a), width = math.sqrt((x1_a-x2_a)**2+(y1_a-y2_a)**2)/5, height = 15, angle=ang_aa, color='black',zorder=7)                  
               
          #       #后车轮
          #       square6 = plt.Rectangle(xy=(x1_a, y1_a), width = math.sqrt((x1_a-x2_a)**2+(y1_a-y2_a)**2)/5, height = -15, angle=ang_aa, color='black',zorder=7)                  
                           
          #       ax.add_patch(square1)
          #       ax.add_patch(square2)
          #       ax.add_patch(square3)
          #       ax.add_patch(square4)
          #       ax.add_patch(square5)
          #       ax.add_patch(square6)










            # 方案2：车身长度固定  
            for i in range(1,4):
                l_x=len(x_a)
                x1_a = x_a[l_x//4*i]
                y1_a = y_a[l_x//4*i]
                
                x2_a = x_a[l_x//4*i+10]
                y2_a = y_a[l_x//4*i+10]
                           
                
                ang_a = angle_of_vector([x2_a-x1_a,y2_a-y1_a],[1,0])
    
                
                if x2_a-x1_a<0 and y2_a-y1_a<=0:
                    x1_a = x2_a - 27* math.cos(ang_a/180*np.pi)
                    y1_a = y2_a + 27* math.sin(ang_a/180*np.pi)
                    
                    #用来遮盖车身上线条的圆形的坐标位置
                    x1_aaa = x2_a - 23.75* math.cos(ang_a/180*np.pi)
                    y1_aaa = y2_a + 23.75* math.sin(ang_a/180*np.pi)
                    x2_aaa = x2_a - 3.25* math.cos(ang_a/180*np.pi)
                    y2_aaa = y2_a + 3.25* math.sin(ang_a/180*np.pi)                      
                    
                    
                    ang_aa = -ang_a
                    ang_a = 180-ang_a 


                    x2_aa = x1_a + 40* math.cos(ang_a/180*np.pi)
                    y2_aa = y1_a - 40* math.sin(ang_a/180*np.pi)
                    ax.arrow(x1_a,y1_a,-x2_aa+x1_a,y2_aa-y1_a,length_includes_head = True,head_width = 13,head_length = 13,fc='royalblue',ec='black',zorder=11)
 

                    
                elif x2_a-x1_a<=0 and y2_a-y1_a>0:
                    x1_a = x2_a - 27* math.cos(ang_a/180*np.pi)
                    y1_a = y2_a - 27* math.sin(ang_a/180*np.pi)
                    
                    #用来遮盖车身上线条的圆形的坐标位置
                    x1_aaa = x2_a - 23.75* math.cos(ang_a/180*np.pi)
                    y1_aaa = y2_a - 23.75* math.sin(ang_a/180*np.pi)
                    x2_aaa = x2_a - 3.25* math.cos(ang_a/180*np.pi)
                    y2_aaa = y2_a - 3.25* math.sin(ang_a/180*np.pi)                    
                    
                    
                    
                    ang_aa = ang_a
                    ang_a = ang_a-180 
                    
            
                    x2_aa = x1_a + 40* math.cos(ang_a/180*np.pi)
                    y2_aa = y1_a + 40* math.sin(ang_a/180*np.pi)          
                    ax.arrow(x1_a,y1_a,-x2_aa+x1_a,-y2_aa+y1_a,length_includes_head = True,head_width = 13,head_length = 13,fc='royalblue',ec='black',zorder=11)
                   
    
    
                    
                elif x2_a-x1_a>0 and y2_a-y1_a>=0:
                    x1_a = x2_a - 27* math.cos(ang_a/180*np.pi)
                    y1_a = y2_a - 27* math.sin(ang_a/180*np.pi)
                    
                    #用来遮盖车身上线条的圆形的坐标位置
                    x1_aaa = x2_a - 23.75* math.cos(ang_a/180*np.pi)
                    y1_aaa = y2_a - 23.75* math.sin(ang_a/180*np.pi)
                    x2_aaa = x2_a - 3.25* math.cos(ang_a/180*np.pi)
                    y2_aaa = y2_a - 3.25* math.sin(ang_a/180*np.pi)                      
                    
                    
                    ang_aa = ang_a
                    ang_a = ang_a-180 


                    x2_aa = x1_a + 40* math.cos(ang_a/180*np.pi)
                    y2_aa = y1_a + 40* math.sin(ang_a/180*np.pi)          
                    ax.arrow(x1_a,y1_a,-x2_aa+x1_a,-y2_aa+y1_a,length_includes_head = True,head_width = 13,head_length = 13,fc='royalblue',ec='black',zorder=11)
                        
    

    
                    
                elif x2_a-x1_a>=0 and y2_a-y1_a<0:
                    x1_a = x2_a - 27* math.cos(ang_a/180*np.pi)
                    y1_a = y2_a + 27* math.sin(ang_a/180*np.pi)
                    
                    #用来遮盖车身上线条的圆形的坐标位置
                    x1_aaa = x2_a - 23.75* math.cos(ang_a/180*np.pi)
                    y1_aaa = y2_a + 23.75* math.sin(ang_a/180*np.pi)
                    x2_aaa = x2_a - 3.25* math.cos(ang_a/180*np.pi)
                    y2_aaa = y2_a + 3.25* math.sin(ang_a/180*np.pi)                      
                    
                    
                    ang_aa = -ang_a
                    ang_a = 180-ang_a    


                    x2_aa = x1_a + 40* math.cos(ang_a/180*np.pi)
                    y2_aa = y1_a - 40* math.sin(ang_a/180*np.pi)          
                    # ax.arrow(x1_a,y1_a,-x2_aa+x1_a,y2_aa-y1_a,length_includes_head = True,head_width = 13,head_length = 13,fc='darkorange',ec='black',zorder=9)
                    ax.arrow(x1_a,y1_a,-x2_aa+x1_a,y2_aa-y1_a,length_includes_head = True,head_width = 13,head_length = 13,fc='royalblue',ec='black',zorder=11)  
                    
                       
                          
    
                # #车身(长度用前后步数差的距离衡量)                          
                # square1 = plt.Rectangle(xy=(x2_a, y2_a), width = math.sqrt((x1_a-x2_a)**2+(y1_a-y2_a)**2), height = 10, angle=ang_a, color='royalblue',zorder=8)                  
               
                #车身 （长度固定）
                square1 = plt.Rectangle(xy=(x2_a, y2_a), width = 27, height = 7, angle=ang_a, color='royalblue',ec='black',zorder=12)                  
                              
                #车身
                square2 = plt.Rectangle(xy=(x2_a, y2_a), width = 27, height = -7, angle=ang_a, color='royalblue',ec='black',zorder=12)                  
    
                #前车轮                          
                square3 = plt.Rectangle(xy=(x2_a, y2_a), width = 5, height = 10, angle=ang_a, color='black',ec='black',zorder=11)                  
               
                #前车轮
                square4 = plt.Rectangle(xy=(x2_a, y2_a), width = 5, height = -10, angle=ang_a, color='black',ec='black',zorder=11)                  
     
                #后车轮                          
                square5 = plt.Rectangle(xy=(x1_a, y1_a), width = 5, height = 10, angle=ang_aa, color='black',ec='black',zorder=11)                  
               
                #后车轮
                square6 = plt.Rectangle(xy=(x1_a, y1_a), width = 5, height = -10, angle=ang_aa, color='black',ec='black',zorder=11)                  



                
                
                #画两个圆形用来遮挡车身上的黑色竖线 
               
                cir_ = Circle(xy = (x1_aaa, y1_aaa), radius=3.25, facecolor= 'royalblue', zorder=13)
                cir__ = Circle(xy = (x2_aaa, y2_aaa), radius=3.25, facecolor= 'royalblue', zorder=13)
                ax.add_patch(cir_)
                ax.add_patch(cir__)

                           
                ax.add_patch(square1)
                ax.add_patch(square2)
                ax.add_patch(square3)
                ax.add_patch(square4)
                ax.add_patch(square5)
                ax.add_patch(square6)
                ax.add_patch(square6)
               
                
                
                
                #绘制车身上的标号        
                cir_num = Circle(xy = ((x1_a+x2_a)/2, (y1_a+y2_a)/2), radius=7, facecolor= 'white',ec='black', zorder=13)
                ax.add_patch(cir_num)
                plt.text((x1_a+x2_a)/2, (y1_a+y2_a)/2, str(i), color='black',ha='center',va='center',weight='heavy',fontsize=6,zorder=14)














            #画矩形用来表示小车：B车
            
            
            
            
            # #方案1：车身长度与小车前后的步数长度相关  
            # for i in range(1,4):
            #     l_x=len(x_b)
            #     x1_b = x_b[l_x//4*i]
            #     y1_b = y_b[l_x//4*i]
                
            #     x2_b = x_b[l_x//4*i+10]
            #     y2_b = y_b[l_x//4*i+10]
               
                
            #     ang_b = angle_of_vector([x2_b-x1_b,y2_b-y1_b],[1,0])
    
                
            #     if x2_b-x1_b<0 and y2_b-y1_b<=0:
            #         ang_bb = -ang_b
            #         ang_b = 180-ang_b 

                    
            #     elif x2_b-x1_b<=0 and y2_b-y1_b>0:
            #         ang_bb = ang_b
            #         ang_b = ang_b-180 
    
                    
            #     elif x2_b-x1_b>0 and y2_b-y1_b>=0:
            #         ang_bb = ang_b
            #         ang_b = ang_b-180 
    
                    
            #     elif x2_b-x1_b>=0 and y2_b-y1_b<0:
            #         ang_bb = -ang_b
            #         ang_b = 180-ang_b   
                         
    
                # #车身                           
                # square7 = plt.Rectangle(xy=(x2_b, y2_b), width = math.sqrt((x1_b-x2_b)**2+(y1_b-y2_b)**2), height = 10, angle=ang_b, color='red',zorder=8)                  
               
                # #车身
                # square8 = plt.Rectangle(xy=(x2_b, y2_b), width = math.sqrt((x1_b-x2_b)**2+(y1_b-y2_b)**2), height = -10, angle=ang_b, color='red',zorder=8)                  
    
                # #前车轮                          
                # square9 = plt.Rectangle(xy=(x2_b, y2_b), width = math.sqrt((x1_b-x2_b)**2+(y1_b-y2_b)**2)/5, height = 15, angle=ang_b, color='black',zorder=7)                  
               
                # #前车轮
                # square10 = plt.Rectangle(xy=(x2_b, y2_b), width = math.sqrt((x1_b-x2_b)**2+(y1_b-y2_b)**2)/5, height = -15, angle=ang_b, color='black',zorder=7)                  
     
                # #后车轮                          
                # square11 = plt.Rectangle(xy=(x1_b, y1_b), width = math.sqrt((x1_b-x2_b)**2+(y1_b-y2_b)**2)/5, height = 15, angle=ang_bb, color='black',zorder=7)                  
               
                # #后车轮
                # square12 = plt.Rectangle(xy=(x1_b, y1_b), width = math.sqrt((x1_b-x2_b)**2+(y1_b-y2_b)**2)/5, height = -15, angle=ang_bb, color='black',zorder=7)                  
 
                # ax.add_patch(square7)
                # ax.add_patch(square8)
                # ax.add_patch(square9)
                # ax.add_patch(square10)
                # ax.add_patch(square11)
                # ax.add_patch(square12)





            #方案2：车身长度固定
            for i in range(1,4):
                l_x=len(x_b)
                
                x1_b = x_b[l_x//4*i]
                y1_b = y_b[l_x//4*i] 
                
                x2_b = x_b[l_x//4*i+10]
                y2_b = y_b[l_x//4*i+10]
               
                
                ang_b = angle_of_vector([x2_b-x1_b,y2_b-y1_b],[1,0])
    
                
                if x2_b-x1_b<0 and y2_b-y1_b<=0:
                    x1_b = x2_b - 27* math.cos(ang_b/180*np.pi)
                    y1_b = y2_b + 27* math.sin(ang_b/180*np.pi)
                    
                    
                    #用来遮盖车身上线条的圆形的坐标位置
                    x1_bbb = x2_b - 23.75* math.cos(ang_b/180*np.pi)
                    y1_bbb = y2_b + 23.75* math.sin(ang_b/180*np.pi)
                    x2_bbb = x2_b - 3.25* math.cos(ang_b/180*np.pi)
                    y2_bbb = y2_b + 3.25* math.sin(ang_b/180*np.pi)                      
                    
                    ang_bb = -ang_b
                    ang_b = 180-ang_b 
                    
                  
                    
                    
                    x2_bb = x1_b + 40* math.cos(ang_b/180*np.pi)
                    y2_bb = y1_b - 40* math.sin(ang_b/180*np.pi)
                    ax.arrow(x1_b,y1_b,-x2_bb+x1_b,y2_bb-y1_b,length_includes_head = True,head_width = 13,head_length = 13,fc='red',ec='black',zorder=6)
         

                    
                elif x2_b-x1_b<=0 and y2_b-y1_b>0:
                    x1_b = x2_b - 27* math.cos(ang_b/180*np.pi)
                    y1_b = y2_b - 27* math.sin(ang_b/180*np.pi)
                    
                    #用来遮盖车身上线条的圆形的坐标位置
                    x1_bbb = x2_b - 23.75* math.cos(ang_b/180*np.pi)
                    y1_bbb = y2_b - 23.75* math.sin(ang_b/180*np.pi)                    
                    x2_bbb = x2_b - 3.25* math.cos(ang_b/180*np.pi)
                    y2_bbb = y2_b - 3.25* math.sin(ang_b/180*np.pi)                      
                    
                    ang_bb = ang_b
                    ang_b = ang_b-180 
                    
                    x2_bb = x1_b + 40* math.cos(ang_b/180*np.pi)
                    y2_bb = y1_b + 40* math.sin(ang_b/180*np.pi)          
                    ax.arrow(x1_b,y1_b,-x2_bb+x1_b,-y2_bb+y1_b,length_includes_head = True,head_width = 13,head_length = 13,fc='red',ec='black',zorder=6)
                   
    
                    
                elif x2_b-x1_b>0 and y2_b-y1_b>=0:
                    x1_b = x2_b - 27* math.cos(ang_b/180*np.pi)
                    y1_b = y2_b - 27* math.sin(ang_b/180*np.pi)
                    
                    #用来遮盖车身上线条的圆形的坐标位置
                    x1_bbb = x2_b - 23.75* math.cos(ang_b/180*np.pi)
                    y1_bbb = y2_b - 23.75* math.sin(ang_b/180*np.pi)                    
                    x2_bbb = x2_b - 3.25* math.cos(ang_b/180*np.pi)
                    y2_bbb = y2_b - 3.25* math.sin(ang_b/180*np.pi)                      
                    
                    ang_bb = ang_b
                    ang_b = ang_b-180 
                    
                    
                    x2_bb = x1_b + 40* math.cos(ang_b/180*np.pi)
                    y2_bb = y1_b + 40* math.sin(ang_b/180*np.pi)          
                    ax.arrow(x1_b,y1_b,-x2_bb+x1_b,-y2_bb+y1_b,length_includes_head = True,head_width = 13,head_length = 13,fc='red',ec='black',zorder=6)
                        
                    
    
                    
                elif x2_b-x1_b>=0 and y2_b-y1_b<0:
                    x1_b = x2_b - 27* math.cos(ang_b/180*np.pi)
                    y1_b = y2_b + 27* math.sin(ang_b/180*np.pi)
                    
                    
                    #用来遮盖车身上线条的圆形的坐标位置
                    x1_bbb = x2_b - 23.75* math.cos(ang_b/180*np.pi)
                    y1_bbb = y2_b + 23.75* math.sin(ang_b/180*np.pi)                    
                    x2_bbb = x2_b - 3.25* math.cos(ang_b/180*np.pi)
                    y2_bbb = y2_b + 3.25* math.sin(ang_b/180*np.pi)                     
                    
                    ang_bb = -ang_b
                    ang_b = 180-ang_b   

                    x2_bb = x1_b + 40* math.cos(ang_b/180*np.pi)
                    y2_bb = y1_b - 40* math.sin(ang_b/180*np.pi)          
                    # ax.arrow(x1_b,y1_b,-x2_bb+x1_b,y2_bb-y1_b,length_includes_head = True,head_width = 13,head_length = 13,fc='limegreen',ec='black',zorder=6)
                    ax.arrow(x1_b,y1_b,-x2_bb+x1_b,y2_bb-y1_b,length_includes_head = True,head_width = 13,head_length = 13,fc='red',ec='black',zorder=6)  
                    

                

    
                
                                    
    
                #车身                           
                square7 = plt.Rectangle(xy=(x2_b, y2_b), width = 27, height = 7, angle=ang_b, color='red',ec='black',zorder=8)                  
               
                #车身
                square8 = plt.Rectangle(xy=(x2_b, y2_b), width = 27, height = -7, angle=ang_b, color='red',ec='black',zorder=8)                  
    
                #前车轮                          
                square9 = plt.Rectangle(xy=(x2_b, y2_b), width = 5, height = 10, angle=ang_b, color='black',ec='black',zorder=7)                  
               
                #前车轮
                square10 = plt.Rectangle(xy=(x2_b, y2_b), width = 5, height = -10, angle=ang_b, color='black',ec='black',zorder=7)                  
     
                #后车轮                          
                square11 = plt.Rectangle(xy=(x1_b, y1_b), width = 5, height = 10, angle=ang_bb, color='black',ec='black',zorder=7)                  
               
                #后车轮
                square12 = plt.Rectangle(xy=(x1_b, y1_b), width = 5, height = -10, angle=ang_bb, color='black',ec='black',zorder=7)                  
               
                
                #画两个圆形用来遮挡车身上的黑色竖线 
               
                cir_ = Circle(xy = (x1_bbb, y1_bbb), radius=3.25, facecolor= 'red', zorder=9)
                cir__ = Circle(xy = (x2_bbb, y2_bbb), radius=3.25, facecolor= 'red', zorder=9)
                ax.add_patch(cir_)
                ax.add_patch(cir__)




                                    
                ax.add_patch(square7)
                ax.add_patch(square8)
                ax.add_patch(square9)
                ax.add_patch(square10)
                ax.add_patch(square11)
                ax.add_patch(square12)

                


                #绘制车身上的标号        
                cir_num = Circle(xy = ((x1_b+x2_b)/2, (y1_b+y2_b)/2), radius=7, facecolor= 'white',ec='black', zorder=9)
                ax.add_patch(cir_num)
                plt.text((x1_b+x2_b)/2, (y1_b+y2_b)/2, str(i), color='black',ha='center',va='center',weight='heavy',fontsize=6,zorder=10)
                
                
                #画两个圆用来遮挡车身上的黑色竖线 


            #标注两小车起始点
            cir_a = Circle(xy = (x_a[0], y_a[0]), radius=5, facecolor= 'royalblue',ec='black',zorder=12)
            ax.add_patch(cir_a)

            cir_b = Circle(xy = (x_b[0], y_b[0]), radius=5, facecolor= 'red',ec='black',zorder=11)
            ax.add_patch(cir_b)                  
               
            # #绘制轨迹指示箭头1
            # ax.arrow(x_a[len(x_a)//5],y_a[len(x_a)//5],x_a[len(x_a)//5 +1]-x_a[len(x_a)//5],y_a[len(x_a)//5 +1]-y_a[len(x_a)//5],length_includes_head = True,head_width = 20,head_length = 20,fc='cyan',ec='black',zorder=10)
            # ax.arrow(x_b[len(x_b)//5],y_b[len(x_b)//5],x_b[len(x_b)//5 +1]-x_b[len(x_b)//5],y_b[len(x_b)//5 +1]-y_b[len(x_b)//5],length_includes_head = True,head_width = 20,head_length = 20,fc='cyan',ec='black',zorder=9)
                   

            # #绘制轨迹指示箭头2
            # ax.arrow(x_a[len(x_a)//5*2],y_a[len(x_a)//5*2],x_a[len(x_a)//5*2 +1]-x_a[len(x_a)//5*2],y_a[len(x_a)//5*2 +1]-y_a[len(x_a)//5*2],length_includes_head = True,head_width = 20,head_length = 20,fc='orange',ec='black',zorder=8)
            # ax.arrow(x_b[len(x_b)//5*2],y_b[len(x_b)//5*2],x_b[len(x_b)//5*2 +1]-x_b[len(x_b)//5*2],y_b[len(x_b)//5*2 +1]-y_b[len(x_b)//5*2],length_includes_head = True,head_width = 20,head_length = 20,fc='orange',ec='black',zorder=7)

            
           
            # #绘制轨迹指示箭头3
            # ax.arrow(x_a[len(x_a)//5*3],y_a[len(x_a)//5*3],x_a[len(x_a)//5*3 +1]-x_a[len(x_a)//5*3],y_a[len(x_a)//5*3 +1]-y_a[len(x_a)//5*3],length_includes_head = True,head_width = 20,head_length = 20,fc='limegreen',ec='black',zorder=6)
            # ax.arrow(x_b[len(x_b)//5*3],y_b[len(x_b)//5*3],x_b[len(x_b)//5*3 +1]-x_b[len(x_b)//5*3],y_b[len(x_b)//5*3 +1]-y_b[len(x_b)//5*3],length_includes_head = True,head_width = 20,head_length = 20,fc='limegreen',ec='black',zorder=5)
        
            # #绘制轨迹指示箭头4
            # ax.arrow(x_a[len(x_a)//5*4],y_a[len(x_a)//5*4],x_a[len(x_a)//5*4 +1]-x_a[len(x_a)//5*4],y_a[len(x_a)//5*4 +1]-y_a[len(x_a)//5*4],length_includes_head = True,head_width = 20,head_length = 20,fc='gold',ec='black',zorder=4)
            # ax.arrow(x_b[len(x_b)//5*4],y_b[len(x_b)//5*4],x_b[len(x_b)//5*4 +1]-x_b[len(x_b)//5*4],y_b[len(x_b)//5*4 +1]-y_b[len(x_b)//5*4],length_includes_head = True,head_width = 20,head_length = 20,fc='gold',ec='black',zorder=3)
                   

            
            #   #绘制注解指示箭头            
            # plt.annotate("Obstacle", (o_x[-1],o_y[-1]), xycoords='data',xytext=(o_x[-1], o_y[-1]), arrowprops=dict(arrowstyle='->'),bbox=dict(boxstyle='circle,pad=0.5', fc='yellow', ec='black', lw=1, alpha=0.5))
            # plt.annotate("Initial Point", (x_a[0],y_a[0]), xycoords='data',xytext=(x_a[0], y_a[0]+50), arrowprops=dict(arrowstyle='->'),bbox=dict(boxstyle='round,pad=0.5', fc='yellow', ec='black', lw=1, alpha=0.5))
            # plt.annotate("Initial Point", (x_b[0],y_b[0]), xycoords='data',xytext=(x_b[0], y_b[0]-50), arrowprops=dict(arrowstyle='->'),bbox=dict(boxstyle='round,pad=0.5', fc='yellow', ec='black', lw=1, alpha=0.5))
                           
            


            plt.rcParams['font.sans-serif'] = ['Times New Roman']
            # plt.title('Trajectory of Cars',fontfamily='Times New Roman',fontweight='heavy',fontsize=17)
            
            plt.scatter(x_a[0], y_a[0],color='royalblue',label='Start of Car A',zorder=0)
            plt.plot(x_a, y_a, color='royalblue', label='Trajectory of Car A',linewidth=1.5,zorder=1)
            
            
            plt.scatter(x_b[0], y_b[0],color='red',label='Start of Car B',zorder=1.5)            
            plt.plot(x_b, y_b, color='red', label='Trajectory of Car B',linewidth=1.5, zorder=2)  
            

            plt.scatter(o_x[0], o_y[0],color='black',label='Obstacle')
            
            plt.legend(bbox_to_anchor=(1.05, 0), loc=3, borderaxespad=0,prop={'family' : 'Times New Roman','weight':'heavy'}) # 标签位置
                            
            # plt.legend(loc='lower right',prop={'family' : 'Times New Roman','weight':'heavy'}) # 标签位置
            plt.xlabel('x(m)',fontfamily='Times New Roman',fontweight='heavy',fontsize=12)
            plt.ylabel('y(m)',fontfamily='Times New Roman',fontweight='heavy',fontsize=12)
            plt.yticks(fontproperties = 'Times New Roman', weight='heavy',size = 12)
            plt.xticks(fontproperties = 'Times New Roman', weight='heavy',size = 12)
            plt.grid()
            # plt.xlim(min(min(x_a),min(x_b))-40,max(max(x_a),max(x_b))+40)
            # plt.ylim(min(min(y_a),min(y_b))-40,max(max(y_a),max(y_b))+40)
            # plt.savefig('./graph/paper/Good Result/test_result_d/trajectory.png')           
            plt.savefig('./graph/paper/Good Result/test_result_'+str(j_)+'/Trajectory of Cars.png',dpi=1000)

            plt.show()

         



        
        
        
        # # 分开画图            
        # if episode+1 == Episode:           
            
        #     # 速度对比图                     
        #     plt.figure(2)
        #     # plt.title('Velocity of Cars',fontfamily='Times New Roman',fontweight='heavy',fontsize=17)
        #     plt.plot(Step, V_a, color='royalblue', label='Velocity of Car A')
        #     plt.plot(Step, V_b, color='red', label='Velocity of Car B')
        #     plt.legend(loc='upper right',prop={'family' : 'Times New Roman','weight':'heavy'}) # 标签位置
        #     plt.xlabel('time(step)',fontfamily='Times New Roman',fontweight='heavy',fontsize=12)
        #     plt.ylabel('velocity(m/s)',fontfamily='Times New Roman',fontweight='heavy',fontsize=12)
        #     plt.yticks(fontproperties = 'Times New Roman', weight='heavy',size = 12)
        #     plt.xticks(fontproperties = 'Times New Roman', weight='heavy',size = 12)
        #     plt.grid()
        #     plt.show()            
        #     plt.savefig('./graph/paper/Good Result/test_result_'+str(j_)+'/Velocity of Cars.png',dpi=1000)    
            
            
        #     # 距离对比图
        #     plt.figure(3)
        #     # plt.title('Distance Between Cars',fontfamily='Times New Roman',fontweight='heavy',fontsize=17)
        #     plt.plot(Step, D, color='green', label='Distance Between Car A and Car B')
        #     plt.plot(Step, Line_d, color='fuchsia', label='Reference',linestyle='--')   
        #     plt.fill_between(Step,Line_d,Line_zero,where=(Line_d>Line_zero), facecolor='fuchsia',alpha=0.1)
        #     plt.legend(loc='upper right',prop={'family' : 'Times New Roman','weight':'heavy'}) # 标签位置
        #     plt.xlabel('time(step)',fontfamily='Times New Roman',fontweight='heavy',fontsize=12)
        #     plt.ylabel('distance(m)',fontfamily='Times New Roman',fontweight='heavy',fontsize=12)
        #     plt.yticks(fontproperties = 'Times New Roman', weight='heavy',size = 12)
        #     plt.xticks(fontproperties = 'Times New Roman', weight='heavy',size = 12)
        #     plt.grid()
        #     plt.show()
        #     plt.savefig('./graph/paper/Good Result/test_result_'+str(j_)+'/Distance Between Cars.png',dpi=1000)   

            
        #     # 角度对比图
        #     plt.figure(4)
        #     # plt.title('Angle of Cars',fontfamily='Times New Roman',fontweight='heavy',fontsize=17)
        #     plt.plot(Step, Angle_a, color='royalblue', label='Advantage Angle of Car A')
        #     plt.plot(Step, Angle_b, color='red', label='Advantage Angle of Car B')
        #     plt.plot(Step, Line_a, color='fuchsia', label='Reference',linestyle='--')
        #     plt.fill_between(Step,Line_a,Line_zero,where=(Line_a>Line_zero), facecolor='fuchsia',alpha=0.1)    
        #     plt.legend(loc='upper right',prop={'family' : 'Times New Roman','weight':'heavy'}) # 标签位置
        #     plt.xlabel('time(step)',fontfamily='Times New Roman',fontweight='heavy',fontsize=12)
        #     plt.ylabel('angle(rad)',fontfamily='Times New Roman',fontweight='heavy',fontsize=12)
        #     plt.yticks(fontproperties = 'Times New Roman', weight='heavy',size = 12)
        #     plt.xticks(fontproperties = 'Times New Roman', weight='heavy',size = 12)
        #     plt.grid()
        #     plt.show()
        #     plt.savefig('./graph/paper/Good Result/test_result_'+str(j_)+'/Angle of Cars.png',dpi=1000)  
            

            
        #     # 每回合中A车和B车每一步的量化优势差值对比
        #     step_r_max=[]
        #     for i in range(len(step_r_error)):
        #         step_r_max.append(max(step_r_error))
        #     plt.figure(5)
        #     # plt.title('Error of Total Advantage in Each Step',fontfamily='Times New Roman',fontweight='heavy',fontsize=17)
        #     plt.plot(Step, step_r_error, color='green', label='Total Advantage in Each Step of Car A minus its of Car B')
        #     plt.plot(Step, step_r_0, color='fuchsia', label='Reference',linestyle='--')
        #     plt.fill_between(Step,step_r_0,step_r_max,where=(step_r_max>step_r_0), facecolor='fuchsia',alpha=0.1) 
        #     plt.legend(loc='upper right',prop={'family' : 'Times New Roman','weight':'heavy'}) # 标签位置
        #     plt.xlabel('time(step)',fontfamily='Times New Roman',fontweight='heavy',fontsize=12)
        #     plt.ylabel('error of total advantage',fontfamily='Times New Roman',fontweight='heavy',fontsize=12)
        #     plt.yticks(fontproperties = 'Times New Roman', weight='heavy',size = 12)
        #     plt.xticks(fontproperties = 'Times New Roman', weight='heavy',size = 12)
        #     plt.grid()
        #     plt.show()
        #     plt.savefig('./graph/paper/Good Result/test_result_'+str(j_)+'/Error of Total Advantage in Each Step.png',dpi=1000) 
       
            

        #     # N回合平均综合优势对比图
        #     plt.figure(6)
        #     # plt.title('Average Total Advantage in Each Episode',fontfamily='Times New Roman',fontweight='heavy',fontsize=17)
        #     plt.plot(N, Win_Total_A, color='royalblue', label='Average Total Advantage ration in Each Episode of Car A')
        #     plt.plot(N, Win_Total_B, color='red', label='Average Total Advantage ration in Each Episode of Car B')
        #     plt.plot(N, Line_c, color='fuchsia', label='Reference',linestyle='--')
        #     plt.fill_between(N,Line_c,Line_100,where=(Line_c<Line_100), facecolor='fuchsia',alpha=0.1)          
        #     plt.legend(loc='upper right',prop={'family' : 'Times New Roman','weight':'heavy'}) # 标签位置
        #     plt.xlabel('time(episode)',fontfamily='Times New Roman',fontweight='heavy',fontsize=12)
        #     plt.ylabel('advantage ratio(%)',fontfamily='Times New Roman',fontweight='heavy',fontsize=12)
        #     plt.yticks(fontproperties = 'Times New Roman', weight='heavy',size = 12)
        #     plt.xticks(fontproperties = 'Times New Roman', weight='heavy',size = 12)
        #     plt.grid()
        #     plt.show()
        #     plt.savefig('./graph/paper/Good Result/test_result_'+str(j_)+'/Average Total Advantage in Each Episode.png',dpi=1000) 






        #合起来画图
        if episode+1 == Episode:


            fig,ax=plt.subplots(2,2) 
            plt.figure(11)
           
            
          
            
            # 1.距离对比图
            ax1 = plt.subplot(221) 
            # plt.title('Distance Between Cars',fontfamily='Times New Roman',fontweight='heavy',fontsize=17)
            # plt.title('(b)', y=-2,fontfamily='Times New Roman',fontweight='heavy',fontsize=17)
            ax1.plot(Step, D, color='green', label='Distance Between Car A and Car B')
            ax1.plot(Step, Line_d, color='fuchsia', label='Reference',linestyle='--')    
            plt.fill_between(Step,Line_d,Line_zero,where=(Line_d>Line_zero), facecolor='fuchsia',alpha=0.1)
            plt.legend(loc='upper right',prop={'family' : 'Times New Roman','weight':'heavy'}) # 标签位置
            plt.xlabel('time(step)'+'\n'+'\n'+'(a)',fontfamily='Times New Roman',fontweight='heavy',fontsize=12)
            plt.ylabel('distance(m)',fontfamily='Times New Roman',fontweight='heavy',fontsize=12)
            plt.yticks(fontproperties = 'Times New Roman', weight='heavy',size = 12)
            plt.xticks(fontproperties = 'Times New Roman', weight='heavy',size = 12)
            plt.grid()      
            # plt.tight_layout(3)
            
 
    
           # 2.速度对比图                     
            ax2 = plt.subplot(222) 
            # plt.title('Velocity of Cars',fontfamily='Times New Roman',fontweight='heavy',fontsize=17)
            # plt.title('(a)', y=-2,fontfamily='Times New Roman',fontweight='heavy',fontsize=17)
            ax2.plot(Step, V_a, color='royalblue', label='Velocity of Car A')
            ax2.plot(Step, V_b, color='red', label='Velocity of Car B')        
            plt.legend(loc='upper right',prop={'family' : 'Times New Roman','weight':'heavy'}) # 标签位置
            plt.xlabel('time(step)'+'\n'+'\n'+'(b)',fontfamily='Times New Roman',fontweight='heavy',fontsize=12)
            plt.ylabel('velocity(m/s)',fontfamily='Times New Roman',fontweight='heavy',fontsize=12)
            plt.yticks(fontproperties = 'Times New Roman', weight='heavy',size = 12)
            plt.xticks(fontproperties = 'Times New Roman', weight='heavy',size = 12)
            plt.grid()      
            # plt.tight_layout(3)          

    

            
            # 3.角度对比图
            ax3 = plt.subplot(223) 
            # plt.title('Angle of Cars',fontfamily='Times New Roman',fontweight='heavy',fontsize=17)
            # plt.title('(c)', y=-2,fontfamily='Times New Roman',fontweight='heavy',fontsize=17)
            ax3.plot(Step, Angle_a, color='royalblue', label='Advantage Angle of Car A')
            ax3.plot(Step, Angle_b, color='red', label='Advantage Angle of Car B')
            ax3.plot(Step, Line_a, color='fuchsia', label='Reference',linestyle='--')
            plt.fill_between(Step,Line_a,Line_zero,where=(Line_a>Line_zero), facecolor='fuchsia',alpha=0.1)    
            plt.legend(loc='upper right',prop={'family' : 'Times New Roman','weight':'heavy'}) # 标签位置
            plt.xlabel('time(step)'+'\n'+'\n'+'(c)',fontfamily='Times New Roman',fontweight='heavy',fontsize=12)
            plt.ylabel('angle(rad)',fontfamily='Times New Roman',fontweight='heavy',fontsize=12)
            plt.yticks(fontproperties = 'Times New Roman', weight='heavy',size = 12)
            plt.xticks(fontproperties = 'Times New Roman', weight='heavy',size = 12)
            plt.grid()       
            # plt.tight_layout(3)   
            
            
            
            
            # 4.每回合中A车和B车每一步的量化优势差值对比
            step_r_max=[]
            for i in range(len(step_r_error)):
                step_r_max.append(max(step_r_error))
            ax4 = plt.subplot(224) 
            # plt.title('Error of Total Advantage in Each Step',fontfamily='Times New Roman',fontweight='heavy',fontsize=17)
            # plt.title('(d)', y=-2,fontfamily='Times New Roman',fontweight='heavy',fontsize=17)
            ax4.plot(Step, step_r_error, color='green', label='Total Advantage in Each Step of Car A minus its of Car B')
            ax4.plot(Step, step_r_0, color='fuchsia', label='Reference',linestyle='--')
            plt.fill_between(Step,step_r_0,step_r_max,where=(step_r_max>step_r_0), facecolor='fuchsia',alpha=0.1) 
            plt.legend(loc='upper right',prop={'family' : 'Times New Roman','weight':'heavy'}) # 标签位置
            plt.xlabel('time(step)'+'\n'+'\n'+'(d)',fontfamily='Times New Roman',fontweight='heavy',fontsize=12)
 
            plt.ylabel('error of total advantage',fontfamily='Times New Roman',fontweight='heavy',fontsize=12)
            plt.yticks(fontproperties = 'Times New Roman', weight='heavy',size = 12)
            plt.xticks(fontproperties = 'Times New Roman', weight='heavy',size = 12)
            plt.grid()    
                       
            fig = plt.gcf()
            fig.set_size_inches(16, 9)
            plt.tight_layout(2)
            
                      
            plt.show()
            # plt.savefig('./graph/paper/Good Result/test_result_'+str(j_)+'/Total.png',dpi=1000) 
            plt.savefig('./graph/paper/Good Result/test_result_'+str(j_)+'/Total.png') 
       
            

            # # N回合平均综合优势对比图
            # plt.figure(6)
            # # plt.title('Average Total Advantage in Each Episode',fontfamily='Times New Roman',fontweight='heavy',fontsize=17)
            # plt.plot(N, Win_Total_A, color='royalblue', label='Average Total Advantage ration in Each Episode of Car A')
            # plt.plot(N, Win_Total_B, color='red', label='Average Total Advantage ration in Each Episode of Car B')
            # plt.plot(N, Line_c, color='fuchsia', label='Reference',linestyle='--')
            # plt.fill_between(N,Line_c,Line_100,where=(Line_c<Line_100), facecolor='fuchsia',alpha=0.1)          
            # plt.legend(loc='upper right',prop={'family' : 'Times New Roman','weight':'heavy'}) # 标签位置
            # plt.xlabel('time(episode)',fontfamily='Times New Roman',fontweight='heavy',fontsize=12)
            # plt.ylabel('advantage ratio(%)',fontfamily='Times New Roman',fontweight='heavy',fontsize=12)
            # plt.yticks(fontproperties = 'Times New Roman', weight='heavy',size = 12)
            # plt.xticks(fontproperties = 'Times New Roman', weight='heavy',size = 12)
            # plt.grid()
            # plt.show()
            # plt.savefig('./graph/paper/Good Result/test_result_'+str(j_)+'/Average Total Advantage in Each Episode.png',dpi=1000) 









            
            
            #打印出A车的优势比率
            print("本回合中，A车相对B车处于对抗优势的时间比例为:", '%.2f%%' % (win_total_a*100))            
            #打印出B车的优势比率
            print("本回合中，B车相对A车处于对抗优势的时间比例为:", '%.2f%%' % (win_total_b*100))                      
            #打印出A车和B车的均势比率                   
            print("本回合中，A车相对B车处于综合均势的时间比例为:", '%.2f%%' % (100-win_total_a*100-win_total_b*100))   
   
            print('\n')             
  
            #打印出A车和B车的对抗优势差值 
            print("本回合中，A车相对B车对抗优势差大于0的时间比例为:", '%.2f%%' % ((win_error_1)*100))         
            print("本回合中，A车相对B车对抗优势差小于0的时间比例为:", '%.2f%%' % ((win_error_2)*100))  
            print("本回合中，A车相对B车对抗优势差等于0的时间比例为:", '%.2f%%' % ((win_error_3)*100))  






            #将print的结果输入在txt文件
            with open('./graph/paper/Good Result/test_result_'+str(j_)+'/Advantage.txt', 'w') as f:  # 设置文件对象
          
                #打印出A车的优势比率
                print("本回合中，A车相对B车处于综合优势的时间比例为:", '%.2f%%' % (win_total_a*100),file = f)                              
                #打印出B车的优势比率
                print("本回合中，B车相对A车处于综合优势的时间比例为:", '%.2f%%' % (win_total_b*100),file = f)                
                #打印出A车和B车的均势比率                    
                print("本回合中，A车相对B车处于综合均势的时间比例为:", '%.2f%%' % (100-win_total_a*100-win_total_b*100),file = f)               

                print('\n',file = f)                           
                
                #打印出A车和B车的对抗优势差值                               
                print("本回合中，A车相对B车对抗优势差大于0的时间比例为:", '%.2f%%' % ((win_error_1)*100),file = f)         
                print("本回合中，A车相对B车对抗优势差小于0的时间比例为:", '%.2f%%' % ((win_error_2)*100),file = f)  
                print("本回合中，A车相对B车对抗优势差等于0的时间比例为:", '%.2f%%' % ((win_error_3)*100),file = f)                 
                

          #创建excel文件，保存各项数据          
        if episode+1 == Episode:
            
            #创建excel文件，存放A车和B车的运动轨迹xy坐标数据                
            wb_1 = Workbook() #创建工作簿
            ws_1 = wb_1.active #激活工作表
            ws_1['A1'] = 'A车x坐标'
            ws_1['B1'] = 'A车y坐标'
            ws_1['C1'] = 'B车x坐标'
            ws_1['D1'] = 'B车y坐标' 
            
            for i in range(len(x_a)):                
                ws_1.append([x_a[i],y_a[i],x_b[i],y_b[i]])

            wb_1.save('./graph/paper/Good Result/test_result_'+str(j_)+'/Trajectory or Cars.xlsx')    
            




            #创建excel文件，存放A车和B车的距离数据                
            wb_2 = Workbook() #创建工作簿
            ws_2 = wb_2.active #激活工作表
            ws_2['A1'] = 'Step'
            ws_2['B1'] = 'A车与B车距离D'
            ws_2['C1'] = 'Line_d2'
            ws_2['D1'] = 'Line_d' 
            
            for i in range(len(Step)):                
                ws_2.append([Step[i],D[i],Line_d[i]])

            wb_2.save('./graph/paper/Good Result/test_result_'+str(j_)+'/Distance Between Cars.xlsx') 
            
            


            #创建excel文件，存放A车和B车的速度数据                
            wb_3 = Workbook() #创建工作簿
            ws_3 = wb_3.active #激活工作表
            ws_3['A1'] = 'Step'
            ws_3['B1'] = 'A车速度v_a'
            ws_3['C1'] = 'B车速度v_b'

            
            for i in range(len(Step)):                
                ws_3.append([Step[i],V_a[i],V_b[i]])

            wb_3.save('./graph/paper/Good Result/test_result_'+str(j_)+'/Velocity of Cars.xlsx')       
            
            
            
            #创建excel文件，存放A车和B车的角度数据                
            wb_4 = Workbook() #创建工作簿
            ws_4 = wb_4.active #激活工作表
            ws_4['A1'] = 'Step'
            ws_4['B1'] = 'A车角度angle_a'
            ws_4['C1'] = 'B车角度angle_b'
            ws_4['D1'] = 'Line_a'
            
            for i in range(len(Step)):                
                ws_4.append([Step[i],Angle_a[i],Angle_b[i],Line_a[i]])

            wb_4.save('./graph/paper/Good Result/test_result_'+str(j_)+'/Angle of Cars.xlsx')     
            
            
            
            #创建excel文件，存放每回合每步的综合优势数据                
            wb_5 = Workbook() #创建工作簿
            ws_5 = wb_5.active #激活工作表
            ws_5['A1'] = 'Step'
            ws_5['B1'] = 'A车与B车综合优势差值'
            
            for i in range(len(Step)):                
                ws_5.append([Step[i],step_r_error[i]])       

            wb_5.save('./graph/paper/Good Result/test_result_'+str(j_)+'/Error of Total Advantage in Each Step.xlsx')               
            


            #创建excel文件，存放N回合综合优势比率数据                
            wb_5 = Workbook() #创建工作簿
            ws_5 = wb_5.active #激活工作表
            ws_5['A1'] = 'Episodes'
            ws_5['B1'] = 'A车综合优势比率'
            ws_5['C1'] = 'B车综合优势比率'
            ws_5['D1'] = 'Line_c'
            
            for i in range(len(N)):                
                ws_5.append([N[i],Win_Total_A[i],Win_Total_B[i],Line_c[i]])

            wb_5.save('./graph/paper/Good Result/test_result_'+str(j_)+'/Average Total Advantage in Each Episode.xlsx')               
            

            
        #注意这里除以10**4是为了与前面每一步的优势对应
        R.append(score)

        R__ .append(Episode_advantage_a) 
        R_b__.append(Episode_advantage_b)
        R_error.append((Episode_advantage_a-Episode_advantage_b)/10**3)
        R_0.append(0)
        print("episode:{}, Advantage_A:{}, Advantage_B:{}".format(episode, Episode_advantage_a,Episode_advantage_b))        
        # print("episode:{}, Return_A:{}".format(episode, score))
        env.close()
    print("=======================end=======================")
 
    #累计奖赏曲线
    # plt.figure(7)
    # plt.plot(E,R)
    # plt.legend()
    # plt.show()

    error_episode_1=0
    error_episode_2=0
    error_episode_3=0
    for j in range(len(R_error)):
        if R_error[j]>0:
            error_episode_1+=1
        elif R_error[j]<0:
            error_episode_2+=1            
        elif R_error[j]==0:
            error_episode_3+=1
            
    ppp_1 = error_episode_1/(error_episode_1+error_episode_2+error_episode_3)
    ppp_2 = error_episode_2/(error_episode_1+error_episode_2+error_episode_3)
    ppp_3 = error_episode_3/(error_episode_1+error_episode_2+error_episode_3)
    
    print(str(len(R_error))+"回合中，A车相对B车对抗优势的差值大于0的时间比例为:", '%.2f%%' % (ppp_1*100))
    print(str(len(R_error))+"回合中，A车相对B车对抗优势的差值小于0的时间比例为:", '%.2f%%' % (ppp_2*100))                          
    print(str(len(R_error))+"回合中，A车相对B车对抗优势的差值等于0的时间比例为:", '%.2f%%' % (ppp_3*100))                      


    #写入TXT
    f_ = open('./graph/paper/Good Result/test_result_'+str(j_)+'/Advantage.txt', 'a+')
    print('\n',file=f_) 
    print(str(len(R_error))+"回合中，A车相对B车对抗优势的差值大于0的时间比例为:", '%.2f%%' % (ppp_1*100),file = f_)
    print(str(len(R_error))+"回合中，A车相对B车对抗优势的差值小于0的时间比例为:", '%.2f%%' % (ppp_2*100),file = f_)                          
    print(str(len(R_error))+"回合中，A车相对B车对抗优势的差值等于0的时间比例为:", '%.2f%%' % (ppp_3*100),file = f_)                      
        
    R_error_max=[]    
    for i in range(len(R_error)):
        R_error_max.append(max(R_error))
        
    
    plt.figure(7)
    # plt.title('Error of Total Advantage in Each Episode',fontfamily='Times New Roman',fontweight='heavy',fontsize=17)
    plt.plot(E, R_error, color='green', label='Total Advantage in Each Episode of Car A minus its of Car B')
    plt.plot(E, R_0, color='fuchsia', label='Reference',linestyle='--')
    plt.fill_between(E,R_0,R_error_max,where=(R_error_max>R_0), facecolor='fuchsia',alpha=0.1) 
    plt.legend(loc='upper right',prop={'family' : 'Times New Roman','weight':'heavy'}) # 标签位置
    plt.xlabel('time(episode)',fontfamily='Times New Roman',fontweight='heavy',fontsize=12)
    plt.ylabel('error of total advantage',fontfamily='Times New Roman',fontweight='heavy',fontsize=12)
    plt.yticks(fontproperties = 'Times New Roman', weight='heavy',size = 12)
    plt.xticks(fontproperties = 'Times New Roman', weight='heavy',size = 12)
    plt.grid()
    plt.show()
    plt.savefig('./graph/paper/Good Result/test_result_'+str(j_)+'/Error of Total Advantage in Each Episode.png',dpi=1000) 

    #创建excel文件，存放所有回合综合优势数据                
    wb_6 = Workbook() #创建工作簿
    ws_6 = wb_6.active #激活工作表
    ws_6['A1'] = 'Episodes'
    ws_6['B1'] = 'A车与B车综合优势的差值'

    
    for i in range(len(E)):                
        ws_6.append([E[i],R_error[i]])

    wb_6.save('./graph/paper/Good Result/test_result_'+str(j_)+'/Error of Total Advantage in Each Episode.xlsx')               


    # env and RL param
    
Episode = 1
step = 400
    
    
torch.set_num_threads(config['num_threads'])
os.environ['MKL_NUM_THREADS'] = str(config['num_threads'])
os.environ["KMP_DUPLICATE_LIB_OK"]="TRUE"
if not config['random_seed']:
    torch.cuda.manual_seed_all(config['seed'])
    torch.manual_seed(config['seed'])
    np.random.seed(config['seed'])
    random.seed(config['seed'])
        
    
# method
env = CarCombatEnv()
sac = SAC(env)
sac.load_nets('./',19999)

run(env,sac)